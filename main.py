import openpyxl
from pathlib import Path
from dataclasses import dataclass
from typing import Optional, Dict, List

@dataclass
class IPConfig:
    original_name: str
    role: str  # 'master' or 'slave'
    read_write: str
    original_bit_width: int
    original_frequency: int
    original_clk_domain: str
    final_bit_width: int = 0
    final_frequency: int = 0
    final_protocol: str = "-"
    final_clk_domain: str = "-"
    connected_interconnect: str = "-"

@dataclass
class InterconnectConfig:
    name: str
    bit_width: int
    frequency: int
    protocol: str
    clk_domain: str
    master_ips: List[str]  # Original IP names (ip1, ip2, etc.)
    slave_ips: List[str]   # Original IP names (ip3, ip4, etc.)

class ConfigGenerator:
    def __init__(self):
        self.ip_configs: Dict[str, IPConfig] = {}  # Key: M1/S1 format
        self.original_ip_map: Dict[str, str] = {}  # Original name to M1/S1 mapping
        self.interconnect_configs: Dict[str, InterconnectConfig] = {}
        self.script_dir = Path(__file__).parent
        self.master_count = 0
        self.slave_count = 0
    
    def find_excel_file(self) -> Path:
        """Finds the first .xls or .xlsx file in script directory"""
        for file in self.script_dir.glob("*.*"):
            if file.suffix.lower() in ('.xls', '.xlsx'):
                return file
        raise FileNotFoundError("No Excel file found in script directory")
    
    def read_excel(self) -> None:
        """Reads the Excel file with two sheets (IPs and Interconnects)"""
        excel_file = self.find_excel_file()
        print(f"\n[DEBUG] Found Excel file: {excel_file.name}")
        
        try:
            workbook = openpyxl.load_workbook(excel_file)
            
            # =============================================
            # Process IP sheet (first sheet) - just store basic info
            # =============================================
            print("\n[DEBUG] Processing IP Sheet:")
            ip_sheet = workbook.worksheets[0]
            for row_idx, row in enumerate(ip_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip if IP name is empty
                    continue
                
                original_name = str(row[0]).strip()
                
                # Store original IP info (role will be determined from Sheet2)
                self.original_ip_map[original_name] = original_name  # Temporary mapping
                print(f"[DEBUG] Found IP: {original_name}")
            
            # =============================================
            # Process Interconnect sheet (second sheet)
            # =============================================
            if len(workbook.worksheets) > 1:
                print("\n[DEBUG] Processing Interconnect Sheet:")
                interconnect_sheet = workbook.worksheets[1]
                
                # Find column indices for key columns
                header_row = next(interconnect_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                try:
                    ic_name_col = header_row.index("interconnect name")
                    masters_col = header_row.index("set of masters")
                    slaves_col = header_row.index("set of slaves")
                except ValueError as e:
                    raise ValueError("Could not find required columns in Sheet2") from e
                
                for row_idx, row in enumerate(interconnect_sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[ic_name_col]:  # Skip if interconnect name is empty
                        continue
                    
                    ic_name = str(row[ic_name_col]).strip()
                    
                    # Get masters and slaves for this interconnect
                    master_ips = []
                    if row[masters_col]:
                        master_ips = [ip.strip() for ip in str(row[masters_col]).split(',')]
                    
                    slave_ips = []
                    if row[slaves_col]:
                        slave_ips = [ip.strip() for ip in str(row[slaves_col]).split(',')]
                    
                    print(f"[DEBUG] Interconnect {ic_name} has masters: {master_ips} and slaves: {slave_ips}")
                    
                    interconnect = InterconnectConfig(
                        name=ic_name,
                        bit_width=int(row[1]),
                        frequency=int(row[2]),
                        protocol=str(row[3]),
                        clk_domain=str(row[4]),
                        master_ips=master_ips,
                        slave_ips=slave_ips
                    )
                    self.interconnect_configs[interconnect.name] = interconnect
            
            # =============================================
            # Now properly identify masters and slaves
            # =============================================
            self._identify_masters_slaves()
            
            # =============================================
            # Apply interconnect properties to IPs
            # =============================================
            self._apply_interconnect_properties()
            
            # Debug print final IP configurations
            print("\n[DEBUG] Final IP Configurations:")
            for ip_name, config in sorted(self.ip_configs.items()):
                print(f"{ip_name} ({config.role.upper()}): Connected to {config.connected_interconnect}")
            
            print(f"\nProcessed {len(self.ip_configs)} IPs and {len(self.interconnect_configs)} interconnects")
            
        except Exception as e:
            print(f"\n[ERROR] Reading Excel file: {str(e)}")
            raise
    
    def _identify_masters_slaves(self):
        """Identifies masters and slaves based on Sheet2 data"""
        print("\n[DEBUG] Identifying masters and slaves:")
        
        # First pass: Identify all masters from "Set of Masters" columns
        all_masters = set()
        for ic_config in self.interconnect_configs.values():
            all_masters.update(ic_config.master_ips)
        
        # Second pass: Identify all slaves from "Set of Slaves" columns
        all_slaves = set()
        for ic_config in self.interconnect_configs.values():
            all_slaves.update(ic_config.slave_ips)
        
        # Check for IPs listed as both master and slave
        conflict_ips = all_masters.intersection(all_slaves)
        if conflict_ips:
            print(f"[WARNING] IPs listed as both master and slave: {conflict_ips}")
        
        # Create M1/S1 names and IP configurations
        self.master_count = 0
        self.slave_count = 0
        self.original_ip_map = {}
        
        # Process masters first
        for master_ip in sorted(all_masters):
            self.master_count += 1
            ip_name = f"M{self.master_count}"
            self.original_ip_map[master_ip] = ip_name
            
            # Create IP config (basic info will be updated when we process Sheet1 again)
            self.ip_configs[ip_name] = IPConfig(
                original_name=master_ip,
                role='master',
                read_write='-',
                original_bit_width=0,
                original_frequency=0,
                original_clk_domain='-'
            )
            print(f"[DEBUG] Identified master: {master_ip} -> {ip_name}")
        
        # Process slaves
        for slave_ip in sorted(all_slaves):
            self.slave_count += 1
            ip_name = f"S{self.slave_count}"
            self.original_ip_map[slave_ip] = ip_name
            
            # Create IP config (basic info will be updated when we process Sheet1 again)
            self.ip_configs[ip_name] = IPConfig(
                original_name=slave_ip,
                role='slave',
                read_write='-',
                original_bit_width=0,
                original_frequency=0,
                original_clk_domain='-'
            )
            print(f"[DEBUG] Identified slave: {slave_ip} -> {ip_name}")
        
        # Now process Sheet1 again to fill in the IP details
        workbook = openpyxl.load_workbook(self.find_excel_file())
        ip_sheet = workbook.worksheets[0]
        
        for row in ip_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            original_name = str(row[0]).strip()
            if original_name in self.original_ip_map:
                ip_name = self.original_ip_map[original_name]
                ip_config = self.ip_configs[ip_name]
                
                # Update IP details from Sheet1
                ip_config.read_write = str(row[1])
                ip_config.original_bit_width = int(row[2])
                ip_config.original_frequency = int(row[3])
                ip_config.original_clk_domain = str(row[4])
                
                # Initialize final values with original values
                ip_config.final_bit_width = ip_config.original_bit_width
                ip_config.final_frequency = ip_config.original_frequency
                ip_config.final_clk_domain = ip_config.original_clk_domain
    
    def _apply_interconnect_properties(self):
        """Updates IP properties based on connected interconnects"""
        print("\n[DEBUG] Applying interconnect properties:")
        
        for ic_name, ic_config in self.interconnect_configs.items():
            # Process masters for this interconnect
            for original_ip in ic_config.master_ips:
                if original_ip in self.original_ip_map:
                    ip_name = self.original_ip_map[original_ip]
                    ip_config = self.ip_configs[ip_name]
                    ip_config.connected_interconnect = ic_name
                    
                    # Update properties from interconnect
                    ip_config.final_bit_width = ic_config.bit_width
                    ip_config.final_frequency = ic_config.frequency
                    ip_config.final_protocol = ic_config.protocol
                    ip_config.final_clk_domain = ic_config.clk_domain
                    
                    print(f"[DEBUG] Updated master {ip_name} with {ic_name} properties")
                else:
                    print(f"[WARNING] Master IP {original_ip} not found in original IP list")
            
            # Process slaves for this interconnect
            for original_ip in ic_config.slave_ips:
                if original_ip in self.original_ip_map:
                    ip_name = self.original_ip_map[original_ip]
                    ip_config = self.ip_configs[ip_name]
                    ip_config.connected_interconnect = ic_name
                    
                    # Update properties from interconnect
                    ip_config.final_bit_width = ic_config.bit_width
                    ip_config.final_frequency = ic_config.frequency
                    ip_config.final_protocol = ic_config.protocol
                    ip_config.final_clk_domain = ic_config.clk_domain
                    
                    print(f"[DEBUG] Updated slave {ip_name} with {ic_name} properties")
                else:
                    print(f"[WARNING] Slave IP {original_ip} not found in original IP list")
    
    def generate_config_file(self) -> None:
        """Generates config.txt in the same directory"""
        output_file = self.script_dir / "config.txt"
        
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                # Write header
                header = [
                    "IP NAME".ljust(15),
                    "TYPE".ljust(10),
                    "READ/WRITE".ljust(15),
                    "BIT WIDTH".ljust(15),
                    "FREQUENCY".ljust(15),
                    "PROTOCOL".ljust(15),
                    "CLK DOMAIN".ljust(15),
                    "INTERCONNECT".ljust(15),
                    "ORIGINAL IP".ljust(15)
                ]
                f.write("".join(header) + "\n")
                f.write("=" * 150 + "\n")
                
                # Write IP configurations
                for ip_name, config in sorted(self.ip_configs.items()):
                    row = [
                        ip_name.ljust(15),
                        ("MASTER" if config.role == 'master' else "SLAVE").ljust(10),
                        config.read_write.ljust(15),
                        str(config.final_bit_width).ljust(15),
                        str(config.final_frequency).ljust(15),
                        config.final_protocol.ljust(15),
                        config.final_clk_domain.ljust(15),
                        config.connected_interconnect.ljust(15),
                        config.original_name.ljust(15)
                    ]
                    f.write("".join(row) + "\n")
            
            print(f"\nConfig file generated: {output_file.name}")
            
            # Print final output to console
            print("\nFinal Output Preview:")
            print("".join(header))
            print("-" * 150)
            for ip_name, config in sorted(self.ip_configs.items()):
                row = [
                    ip_name.ljust(15),
                    ("MASTER" if config.role == 'master' else "SLAVE").ljust(10),
                    config.read_write.ljust(15),
                    str(config.final_bit_width).ljust(15),
                    str(config.final_frequency).ljust(15),
                    config.final_protocol.ljust(15),
                    config.final_clk_domain.ljust(15),
                    config.connected_interconnect.ljust(15),
                    config.original_name.ljust(15)
                ]
                print("".join(row))
            
        except Exception as e:
            print(f"\n[ERROR] Generating config file: {str(e)}")
            raise

if __name__ == "__main__":
    print("IP-Interconnect Config Generator")
    print("=" * 60)
    
    generator = ConfigGenerator()
    try:
        generator.read_excel()
        generator.generate_config_file()
        print("\nDone! Check config.txt in the same folder")
    except Exception as e:
        print(f"\nError occurred: {str(e)}")
